from django.contrib.auth.models import Group
from django.db import models
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.generic import UpdateView, ListView, CreateView, DeleteView, View
from django.http import HttpResponseForbidden,HttpResponse
from user_account.models import UserNet
from user_account.permissions import AdminPermissionMixin
from .forms import *
from .models import *
from journal.models import *
from .utils import TimetableSettigns, CurruculumMixin
from institutions.permissions import InstitutionsMixin
from modules.users import generate_login
from modules.weeks import get_dates
import random
import re
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.db.models import Q
from journal.utils import Journal
from django.contrib.auth.models import Group
from io import BytesIO
from openpyxl import load_workbook
import tablib





class ClassView(PermissionRequiredMixin, CreateView):
    form_class = ClassForm
    template_name = 'classes/class.html'
    permission_required = 'classes.view_classes'
    def get_success_url(self):
        return reverse('Class')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Классы'
        context['classes'] = Classes.objects.filter(
                institution=self.request.user.institution.pk,year=self.request.user.institution.year.pk)
        return context

    def get_form_kwargs(self):
        kwargs = super(ClassView, self).get_form_kwargs()
        kwargs['edit'] = True
        kwargs['teacher']=self.request.user.institution
        return kwargs


    def form_valid(self, form):
        form.instance.institution_id = self.request.user.institution.pk
        form.instance.year_id = self.request.user.institution.year.pk
        return super().form_valid(form)

class GroupView(PermissionRequiredMixin, CreateView):
    form_class = GroupForm
    template_name = 'classes/groups.html'
    permission_required = 'classes.view_classes'
    def get_success_url(self):
        return reverse('Groups')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Классы'
        context['classes'] = Classes.objects.filter(
                institution=self.request.user.institution.pk,year=self.request.user.institution.year.pk)
        return context

    def get_form_kwargs(self):
        kwargs = super(GroupView, self).get_form_kwargs()
        kwargs['edit'] = True
        kwargs['teacher']=self.request.user.institution
        return kwargs


    def form_valid(self, form):
        form.instance.institution_id = self.request.user.institution.pk
        form.instance.year_id = self.request.user.institution.year.pk
        return super().form_valid(form)



def class_edit_view(request,pk):
    class_info=Classes.objects.get(pk=pk)
    context={}
    if class_info.institution.pk==request.user.institution.pk:
        if request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            context['title'] = 'Редактирование группы'
            info_form=GroupForm(instance=class_info,teacher=request.user.institution)
            
        else:
            context['title'] = 'Редактирование класа'
            info_form=ClassForm(instance=class_info,teacher=request.user.institution)
        subgroups_form=SubgroupsForm(class_number=class_info.class_number, profile=class_info.сurriculum)
        if request.method=="POST":

            if 'info' in request.POST:

                
                if request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
                    
                    in_form=GroupForm(request.POST,instance=class_info,teacher=request.user.institution)
                else:
                    in_form=ClassForm(request.POST,instance=class_info,teacher=request.user.institution)
                if in_form.is_valid():
                    in_form.save()
                    messages.success(request, 'Информация обновлена!')
                    return redirect(request.META.get('HTTP_REFERER'))
            if 'subroups' in request.POST:
                sub_form=SubgroupsForm(request.POST,class_number=class_info.class_number, profile=class_info.сurriculum)
                sub_form.instance.class_pk=class_info
                if sub_form.is_valid():
                    sub_form.save()
                    return redirect(request.META.get('HTTP_REFERER'))
        class_subgroups=Subgroups.objects.filter(class_pk=pk)
        context['form'] = info_form
        context['subroups'] = subgroups_form
        context['subgroups'] = class_subgroups
        return render(request,'classes/class_edit.html',context)
    
    else:
        return HttpResponseForbidden()

class SubgroupView(View,Journal):
    form_class = ClassForm
    template_name = 'classes/subgroup.html'
    def get_classes(self):
        

        if not self.request.user.has_perm('classes.view_classes'):
            return Classes.objects.filter(class_teacher=self.request.user, year=self.request.user.institution.year.pk).order_by('class_number','letter')
        else:
           
            return Classes.objects.filter(institution=self.request.user.institution, year=self.request.user.institution.year.pk).order_by('class_number','letter')
    def get_student(self):
        return Student.objects.filter(class_pk=self.get_class(),user__isnull=False,user__is_active=True).order_by('user')
    def get_class(self):


        try:
            if not self.request.user.has_perm('classes.view_classes'):

                class_info=self.get_classes().first()
            elif 'pk' in self.kwargs:
                class_info=Classes.objects.get(pk=self.kwargs['pk'])
            else:
                class_info=Classes.objects.filter(institution=self.request.user.institution,year=self.request.user.institution.year.pk).first()
        except Classes.DoesNotExist:
            class_info=None

        return class_info
    def get_subgroups(self):
        return Subgroups.objects.filter(class_pk=self.get_class()) or None


    def get_context(self):
        context = {}
        if self.request.user.institution.typeInstitutions.title != 'Профессиональная образовательная организация' and self.get_class():
   
            context['title'] = 'Подгруппы '+str(self.get_class().class_number)+str(self.get_class().letter)+' класса'

        elif self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация' and self.get_class():
            context['title'] = 'Подгруппы '+ str(self.get_class()) +' группы'
        else:
            context['title'] = 'Подгруппы'
        context['classes']=self.get_classes()
        context['students']=self.get_student()
        context['class']=self.get_class()
        context['subgroup_list']=self.get_subgroups()
        if self.get_subgroups():
            arr = []
            sb = None
            for subroup in self.get_subgroups():
                
                if sb != None:
                    if subroup.subject_pk.pk!=sb.pk:
                        arr.append(subroup.subject_pk)
                else:
                    arr.append(subroup.subject_pk)
                sb = subroup.subject_pk
                
            context['subgroups']=arr
        return context

    def get(self, request, *args, **kwargs):
        
        return render(request, self.template_name,self.get_context())

    def post(self, request, *args, **kwargs):
        
        if self.get_subgroups():
            for subgroup in self.get_subgroups().distinct():
                subgroup_pk=subgroup.subject_pk.pk
                for student in self.get_student():
                    student_pk=student.pk
                    sub=self.request.POST.get("sub"+str(subgroup_pk)+str(student_pk))
                    if sub:
                        select_sub=Subgroups.objects.get(pk=sub)
                        check_sub=StudentSubgroup.objects.filter(student=student_pk, subject=select_sub.subject_pk )
                        if check_sub:
                            
                            check_sub.update(subgroup=  sub )
                        else:
                            sub_save=StudentSubgroup.objects.create(student=student,  subgroup_id=sub,subject=select_sub.subject_pk)

        return redirect(request.META.get('HTTP_REFERER'))

class DeleteClassView(PermissionRequiredMixin,DeleteView):

    permission_required='classes.delete_classes'
    def get_object(self, **kwargs):
        id_ = int(self.request.POST.get("pk"))
        return get_object_or_404(Classes, id=id_)

    def get_success_url(self):
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            return reverse('Groups')
        else:
            return reverse('Class')




class СurriculumView(PermissionRequiredMixin,ListView):
    model = Сurriculum
    template_name = 'classes/curriculum.html'
    permission_required = 'classes.view_сurriculum'
    def get_context_data(self):
        context = super().get_context_data()
        context['title'] = 'Учебные планы'
        context['curriculums'] = Сurriculum.objects.filter(
            institution=self.request.user.institution.pk,year=self.request.user.institution.year.pk)
        return context

class СurriculumCreateView(PermissionRequiredMixin, CurruculumMixin, CreateView):
    form_class = СurriculumForm 
    
    permission_required = 'classes.add_сurriculum'

    def get_template_names(self):
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            return ['classes/poo_curriculum.html']
        else:
            return ['classes/add_curriculum.html']
    def get_context_data(self):

        context=super().get_context_data()
        context['title']='Добавление учебного плана'
        c_def=self.get_curriculum_context(title='Добавление учебного плана')
        return dict(list(context.items())+list(c_def.items()))

    def form_valid(self, form):
        form.instance.institution_id = self.request.user.institution.pk
        form.instance.year_id = self.request.user.institution.year.pk
        return super().form_valid(form)

    def get_success_url(self):
        self.form_save()
        return reverse('Curriculum')

class СurriculumUpdateView(PermissionRequiredMixin, CurruculumMixin, UpdateView):
    model=Сurriculum
    form_class = СurriculumForm

    permission_required = 'classes.change_сurriculum'
    def get_template_names(self):
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            return ['classes/poo_curriculum.html']
        else:
            return ['classes/add_curriculum.html']
    def get_context_data(self):

        context=super().get_context_data()
        context['title']='Редактирование учебного плана'
        c_def=self.get_curriculum_context(title='Редактирование учебного плана')
        return dict(list(context.items())+list(c_def.items()))

    def form_valid(self, form):
        form.instance.institution_id = self.request.user.institution.pk
        form.instance.year_id = self.request.user.institution.year.pk
        return super().form_valid(form)

    def get_success_url(self):
        self.form_save()
        return reverse('Curriculum')
class DeleteCurriculum(InstitutionsMixin,AdminPermissionMixin, DeleteView):
    template_name = 'institutions/curriculum.html'

    def get_object(self, **kwargs):
        id_ = self.kwargs.get("pk")
        return get_object_or_404(Сurriculum, id=id_)

    def get_success_url(self):
        return reverse('Curriculum')

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)



class LoadView(PermissionRequiredMixin, TimetableSettigns,SuccessMessageMixin, View):
    form_class = СurriculumForm
    template_name = 'classes/load.html'
    permission_required = 'classes.view_load'


    def get_success_url(self):
        return reverse('Load')

    def get(self, request, *args, **kwargs):
        context = {}
        context['subjects'] = self.get_info()
        context['teachers']=UserNet.objects.filter(institution=self.request.user.institution.pk,groups__name__in=['Учитель','Преподаватель'],is_active=True).distinct()
        context['classes']=Classes.objects.filter(institution=self.request.user.institution.pk, year=self.request.user.institution.year.pk)
        context['class']=self.get_class()
        context['title'] = 'Учебная нагрузка'
        return render(request, self.template_name,context)
    def post(self, request, *args, **kwargs):
        subjects_list=self.get_info()
        class_pk=self.get_class().pk
        for subject in subjects_list:
            subroups_list=Subgroups.objects.filter(class_pk=class_pk,subject_pk=subject.pk)
            if subroups_list:
                for subgroup in subroups_list:
                    teacher=self.request.POST.get("subject_"+str(subject.pk)+"_subgroup_"+str(subgroup.pk))
                    load_check=Load.objects.filter(class_pk=class_pk,subject_pk=subject,subgroup=subgroup)
                    delete=self.request.POST.get("del_"+str(subject.pk))

                    if delete:
                        for i in load_check: get_load=i.pk
                        get_load=Load.objects.get(pk=get_load).delete()
                    else:

                        if load_check and teacher:
                            for i in load_check: get_teacher=i.pk
                            teacher_get=UserNet.objects.get(pk=teacher)
                            get_load=Load.objects.get(pk=get_teacher)
                            get_load.teacher=teacher_get
                            get_load.save()

                            update_timetable_begin=self.request.POST.get("begin_"+str(subject.pk))
                            update_timetable_end=self.request.POST.get("end_"+str(subject.pk))

                            if update_timetable_begin and update_timetable_end:
                                get_lessons=Lessons.objects.filter(subject_pk=get_load, date__gte=update_timetable_begin, date__lte=update_timetable_end)
                                for lesson in get_lessons:
                                    lesson.teacher=teacher_get
                                    lesson.save()


                        else:
                            if teacher:
                                teacher_load=UserNet.objects.get(pk=teacher)
                                Load.objects.create(class_pk=self.get_class(),subject_pk=subject,subgroup=subgroup,teacher=teacher_load)
            else:
                teacher=self.request.POST.get("subject_"+str(subject.pk))
                load_check=Load.objects.filter(class_pk=class_pk,subject_pk=subject)
                update_timetable_begin=self.request.POST.get("begin_"+str(subject.pk))
                update_timetable_end=self.request.POST.get("end_"+str(subject.pk))
                delete=self.request.POST.get("del_"+str(subject.pk))

                if delete:
                    for i in load_check: get_load=i.pk
                    get_load=Load.objects.get(pk=get_load).delete()
                else:
                    if load_check and teacher:
                        
                        for i in load_check: get_teacher=i.pk
                        teacher_get=UserNet.objects.get(pk=teacher)
                        get_load=Load.objects.get(pk=get_teacher)
                        get_load.teacher=teacher_get
                        get_load.save()
                        if update_timetable_begin and update_timetable_end:
                            get_lessons=Lessons.objects.filter(subject_pk=get_load, date__gte=update_timetable_begin, date__lte=update_timetable_end)
                            for lesson in get_lessons:
                                lesson.teacher=teacher_get
                                lesson.save()

                    else:
                        if teacher:
                            teacher_load=UserNet.objects.get(pk=teacher)
                            Load.objects.create(class_pk=self.get_class(),subject_pk=subject,teacher=teacher_load)


        messages.success(request,'Нагрузка успешно обновлена!')
        return redirect(request.META.get('HTTP_REFERER'))


#Расписание
class Timetable(PermissionRequiredMixin,ListView):
    model = Classes
    template_name = 'classes/timetable_classes.html'
    permission_required = 'classes.view_timetabletemplates'
    def get_context_data(self):
        context = super().get_context_data()
        context['title'] = 'Расписание'
        context['classes'] = Classes.objects.filter(institution=self.request.user.institution.pk, year=self.request.user.institution.year.pk)
        return context

class TimetableTemplatesView(PermissionRequiredMixin,TimetableSettigns,ListView):
    model = TimetableTemplates
    template_name = 'classes/timetable_templates.html'
    permission_required = 'classes.view_timetabletemplates'
    def get_context_data(self):
        context = super().get_context_data()
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            context['title'] = 'Шаблоны расписания '+str(self.get_class())+' группы'
        else:
            context['title'] = 'Шаблоны расписания '+str(self.get_class().class_number)+str(self.get_class().letter)+' класса'
        context['class']=self.get_class()
        context['templates'] = TimetableTemplates.objects.filter(class_pk=self.get_class())
        context['curriculums'] = Сurriculum.objects.filter(
            institution=self.request.user.institution.pk,year=self.request.user.institution.year.pk)
        return context

class AddTimetableTemplate(PermissionRequiredMixin, TimetableSettigns, View):
    template_name = 'classes/timetable_add_template.html'
    permission_required = 'classes.view_timetabletemplates'
    def post(self,request):
        context={}
        context['days'] = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
        context['title'] = 'Добавление шаблона расписания'
        context['loads']=Load.objects.filter(class_pk=self.request.POST.get("class"))
        context['class_pk']=self.request.POST.get("class")
        return render(request, self.template_name,context)

class CreateTimetableTemplate(PermissionRequiredMixin,View):
    permission_required = 'classes.view_timetabletemplates'
    def post(self,request):
        class_pk=request.POST.get("class_pk")
        class_get=Classes.objects.get(pk=class_pk)
        bell_filter=BellTimetable.objects.filter(profile=class_get.bell_profile).order_by('day','lesson')
        profile_name=request.POST.get("profile")
        template=TimetableTemplates.objects.create(сurriculum=class_get.сurriculum,name=profile_name,class_pk=class_get)
        for bell in bell_filter:
            lesson=bell
            loads=request.POST.getlist(str(bell.day)+str((bell.lesson)))
            
            for load in loads:
                if load:
                    get_subject=Load.objects.get(pk=int(load))
                    SubjectTemplate.objects.create(day=bell.day,lesson=bell.lesson,profile=template,subject_pk=get_subject)

        return redirect(class_get.timetable_templates())


class UpdateTimetableTemplate(PermissionRequiredMixin,View):
    template_name = 'classes/timetable_add_template.html'
    permission_required = 'classes.view_timetabletemplates'
    def get_template(self):
        return TimetableTemplates.objects.get(pk=self.kwargs['pk'])
    def get(self, request, *args, **kwargs):
        context={}
        context['class_pk']=self.get_template().class_pk.pk
        context['template']=self.get_template()
        context['loads']=Load.objects.filter(class_pk=self.get_template().class_pk.pk)
        context['days'] = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
        context['title'] = 'Редактирование'+'"'+self.get_template().name+'"'

        return render(request, self.template_name,context)
    def post(self,request,*args, **kwargs):
        class_pk=request.POST.get("class_pk")
        class_get=Classes.objects.get(pk=class_pk)
        bell_filter=BellTimetable.objects.filter(profile=class_get.bell_profile).order_by('day','lesson')
        profile_name=request.POST.get("profile")
        template=TimetableTemplates.objects.get(pk=self.get_template().pk)
        template.name=profile_name
        template.save()

        get_template_subjects=SubjectTemplate.objects.filter(profile=template)
        for subject in get_template_subjects:
            loads=request.POST.getlist(str(subject.day)+str((subject.lesson)))
            if str(subject.subject_pk.pk) not in loads:

                get_template_subject=SubjectTemplate.objects.get(pk=subject.id).delete()

        for bell in bell_filter:
            lesson=bell
            loads=request.POST.getlist(str(bell.day)+str((bell.lesson)))
            new_loads=list(filter(None, loads))
            for load in new_loads:
                 check_template_subjects=SubjectTemplate.objects.filter(day=bell.day,lesson=bell.lesson,profile=template,subject_pk__id=load).values_list('subject_pk')
                 if check_template_subjects:
                    sub_template=check_template_subjects[0][0]
                 else:
                        get_subject=Load.objects.get(pk=int(load))
                        SubjectTemplate.objects.create(day=bell.day,lesson=bell.lesson,profile=template,subject_pk=get_subject)


        return redirect(class_get.timetable_templates())

class TimetableWeek(PermissionRequiredMixin, TimetableSettigns, View):
    template_name = 'classes/timetable_weeks.html'
    permission_required = 'classes.view_timetabletemplates'
    def get(self, request, *args, **kwargs):
        context={}
        context['class_pk']=self.get_class()
        profile=PeriodProfile.objects.get(pk=self.get_class().period_profile.pk)
        if profile.typePeriod==4:
            context['name_period']='Четверть'
        elif profile.typePeriod==3:
            context['name_period']='Триместр'
        elif profile.typePeriod==2:
            context['name_period']='Полугодие'
        if 'period' in self.kwargs:
            context['period_pk']=self.kwargs['period']
        context['class']=self.get_class()
        context['templates']=TimetableTemplates.objects.filter(class_pk=self.get_class())
        context['weeks']=self.get_weeks()
        context['periods']=Periods.objects.filter(profile=self.get_class().period_profile)
        

        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            context['title'] = 'Распределение занятий '+str(self.get_class())+' группы'
        else:
            context['title'] = 'Распределение уроков '+str(self.get_class().class_number)+str(self.get_class().letter)+' класса'
        return render(request, self.template_name,context)
    def post(self, request, *args, **kwargs):
        context={}

        context['class_pk']=self.get_class()
        profile=PeriodProfile.objects.get(pk=self.get_class().period_profile.pk)
        i=0
        for week in self.get_weeks():
            i+=1
            post_week=request.POST.get(str('template')+str(i))
            if post_week:
                start=week[0]
                end=week[1]
                dates=get_dates(start,end)
                get_types = LessonType.objects.get(name='Ответ на уроке')
                for date_week in dates:
                    if date_week>=self.get_period().start and date_week<=self.get_period().end:
                        weekday=date_week.isoweekday()
                        get_lesson=SubjectTemplate.objects.filter(profile__id=post_week,day=weekday)
                        for lesson in get_lesson:
                            lesson_save=Lessons.objects.create(number=lesson.lesson,date=date_week,class_pk=self.get_class(),subject_pk=lesson.subject_pk,teacher=lesson.subject_pk.teacher)
                            lesson_save.types.add(get_types.pk)
        messages.success(request,'Уроки успешно распределены')
        return redirect(request.META.get('HTTP_REFERER'))
class EditLessons(PermissionRequiredMixin,TimetableSettigns,View):
    template_name='classes/timetable_editlesson.html'
    permission_required = 'classes.view_timetabletemplates'
    def get(self,request,class_pk):
        context={}
        
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            context['title']='Редактирование занятий'
        else:
            context['title']='Редактирование уроков'
        context['date']=self.get_date()
        context['lessons']=self.get_lessons()
        context['class']=self.get_class()
        context['loads']=self.get_loads()
        context['teachers']=UserNet.objects.filter(institution=request.user.institution,groups__name__in=['Учитель', 'Преподаватель'],is_active=True)
        return render(request, self.template_name, context)
    def post(self,request,class_pk):
        self.save_edit_timetable()

        return redirect(request.META.get('HTTP_REFERER'))
class DeleteLessons(PermissionRequiredMixin, View):
    permission_required = 'journal.delete_lessons'
    def get(self,request,class_pk):
        begin=request.GET.get("begin")
        end=request.GET.get("end")
        lesson=Lessons.objects.filter(class_pk=class_pk,date__gte=begin,date__lte=end).delete()
        return redirect(request.META.get('HTTP_REFERER'))


#Ученики
class StudentListView(PermissionRequiredMixin, ListView):
    paginate_by = 35
    model = Student
    template_name = 'classes/students.html'
    permission_required = 'classes.view_student'

    def query(self):
        surname=self.request.GET.get('surname')
        name=self.request.GET.get('name')
        middle_name=self.request.GET.get('middle_name')
        class_pk=self.request.GET.get('class_pk')
        if (surname or name or middle_name or class_pk) is not None:
            return surname,name,middle_name,class_pk
    def get_class(self):


        if self.request.user.has_perm('classes.delete_student'):
            return Classes.objects.filter(institution=self.request.user.institution, year=self.request.user.institution.year.pk)
        else:
            return Classes.objects.filter(class_teacher=self.request.user,year=self.request.user.institution.year.pk)
    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            context['title'] = 'Студенты'
        else:
            context['title'] = 'Обучающиеся'
        if self.query():
            context['surname']=self.query()[0]
            context['name']=self.query()[1]
            context['middle_name']=self.query()[2]
            if self.query()[3] == "None":
                context['class_pk']=self.query()[3]
            else:
                context['class_pk']=int(self.query()[3])
        if self.kwargs.get('delete_code'):

            context['delete_code'] = self.kwargs.get('delete_code')

        context['classes']=self.get_class()
        return context

    def get_queryset(self):
        
        if self.query() and (self.query()[0] or self.query()[1] or self.query()[2]):
            users=UserNet.objects.filter(institution=self.request.user.institution, groups__name='Ученик', is_active=True, 
                last_name__icontains=self.query()[0], first_name__icontains=self.query()[1]
                , middle_name__icontains=self.query()[2]

                )

            
        else:
            users=UserNet.objects.filter(institution=self.request.user.institution, groups__name='Ученик', is_active=True)
        if self.query() and self.query()[3] != 'None':
            students=Student.objects.filter(user__in=users,class_pk=int(self.query()[3])).order_by('user__last_name')
        else:
            students=Student.objects.filter(user__in=users, class_pk__in=self.get_class()).order_by('user__last_name')
        return students

class CancelImport(PermissionRequiredMixin,View):

    permission_required='classes.delete_classes'
    def get(self, request, delete_code):
        students = Student.objects.filter(delete_code=delete_code)
        for student in students:
            student.user.delete()
            student.delete()
        return redirect('StudentList')


class ExportStudent(View):

    def get(self,request):
        headers = ('Класс', 'Фамилия','Имя','Отчество' , 'Дата рождения', 'Пол')
        data = []
        data = tablib.Dataset(*data, headers=headers)
        students = Student.objects.filter(user__institution=self.request.user.institution,user__is_active=True)
        for student in students:
            data.append((student.class_pk.class_number,student.user.first_name,student.user.last_name, student.user.middle_name, student.user.birth_day, student.user.gender  ))
        response = HttpResponse(data.xlsx, content_type='application/vnd.ms-excel;charset=utf-8')
        response['Content-Disposition'] = "attachment; filename=export.xlsx"

        return response

class AddStudent(PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    form_class = OldStudentForm
    template_name = 'user_account/user_update.html'
    permission_required = 'classes.add_student'

    def get_form_kwargs(self):
        kwargs = super(AddStudent, self).get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    def form_valid(self, form):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
        self.login = ''
        self.password = ''
        for i in range(8):
            self.password += random.choice(chars)
        for i in range(8):
            self.login += random.choice(chars)
        user = form.save(commit=False)        

        user.code=self.login
        user.institution_id = self.request.user.institution.pk

        user.save()
        group=Group.objects.get(name='Ученик')
        user.groups.add(group)
        Student.objects.create(user=user,class_pk=form.cleaned_data['class'],date_of_enrollment=form.cleaned_data['date_of_enrollment'])
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        if self.request.user.institution.typeInstitutions.title == 'Профессиональная образовательная организация':
            context['title'] = 'Добавление студента'
        else:
            context['title'] = 'Добавление обучающегося'
        return context

    def get_success_url(self):
        return reverse('StudentList')
    def get_success_message(self, cleaned_data):
        ms = f'Пригласительный код: {{ self.login }}'
        success_message = ms
        return success_message % cleaned_data

class ImportStudent(View):
    def get(self,request):
        context={}
        context['title']='Импорт пользователей'
        context['classes'] = self.get_class()
        context['file']=FileTemplates.objects.filter(name='Шаблон для импорта учеников').first()
        return render(request,'classes/import_student.html',context)
    def post(self,request):
        context={}
        context['title']='Результат импорта'
        date_of_enrollment = self.request.POST.get('date')
        class_pk = self.request.POST.get('class_pk')
        my_group=Group.objects.get(name='Ученик')
        class_object = Classes.objects.get(pk=class_pk)
        f = request.FILES['file']
        file_import = load_workbook(filename=BytesIO(f.read()))
        code = generate_login()

        sheet = file_import.active
        rows = sheet.max_row
        cols = sheet.max_column
        for i in range(2,rows+1):

            last_name = sheet.cell(row=i,column=1).value
            first_name = sheet.cell(row=i,column=2).value
            patronymic = sheet.cell(row=i,column=3).value
            gender = sheet.cell(row=i,column=5).value
            birth_day = sheet.cell(row=i,column=4).value
            login = generate_login()
            user_pk=UserNet.objects.create_user(is_active=True,username=login[1],code=login[0],last_name=last_name,first_name=first_name,middle_name=patronymic,gender=gender,birth_day=birth_day,institution=self.request.user.institution)
            Student.objects.create(user=user_pk,class_pk=class_object,date_of_enrollment=date_of_enrollment, delete_code=code)
            my_group.user_set.add(user_pk)
        return redirect('StudentList', delete_code = code)

    def get_class(self):
        if self.request.user.has_perm('classes.delete_student'):
            return Classes.objects.filter(institution=self.request.user.institution, year=self.request.user.institution.year.pk)
        else:
            return Classes.objects.filter(class_teacher=self.request.user,year=self.request.user.institution.year.pk)

class StudentEditView(PermissionRequiredMixin,SuccessMessageMixin,View):
    model = Student
    form_class = StudentUserForm
    second_form_class = StudentForm


    template_name = 'classes/student.html'
    pk_url_kwarg = 'student_pk'
    success_message = 'Информация обновлена'
    error_message='Ошибка'
    permission_required = 'classes.change_student'
    login_url='login'

    def get_student(self):
        return Student.objects.get(pk=self.kwargs['student_pk'])

    def get(self,request, **kwargs):
        context = {}
        
        
        context['title'] = 'Личное дело учащегося'
        if 'form' not in context:
            context['form'] = self.form_class(instance=self.get_student().user)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=self.get_student())

        return render(request, self.template_name, context)
    

    def post(self,request, **kwargs):


        if 'first_name' in request.POST:
            form=self.form_class(request.POST,request.FILES or None,instance=self.get_student().user)
            if form.is_valid():
                form.save()

        if 'class_pk' in request.POST:
            form2=self.second_form_class(request.POST,instance=self.get_student())
            if form2.is_valid():
                form2.save()
            
        return redirect(request.META.get('HTTP_REFERER'))

class Deduction(PermissionRequiredMixin,SuccessMessageMixin,View):
    permission_required = 'classes.delete_student'
    def post(self,request, **kwargs):
        date=self.request.POST.get("date")
        student=self.request.POST.get("pk")
        get_student=Student.objects.get(pk=student)
        get_student.user.is_active=False
        get_student.user.save()
        StudentShifting.objects.create(student=get_student, type_shift=2, date=date)
        return redirect(request.META.get('HTTP_REFERER'))

class ReturnStudents(View):

    def get(self,request,class_pk):

        
        students_get=Student.objects.filter(class_pk=class_pk)
        
        students_fio=[]
        students_pk=[]

        for student in students_get:
            students_fio.append(student.user.last_name+ ' ' +student.user.first_name)
            students_pk.append(student.pk)
        response={
            'fio':students_fio,
            'pk': students_pk  
        }
        return JsonResponse(response)

